import rpy2.robjects as robjects
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from rpy2.robjects.packages import importr
import os
import pandas as pd
import re
import numpy as np
import math
import scipy
from openpyxl.utils import get_column_letter
from collections import Counter
from scipy.fftpack import fft

#Calculate the inter-amino features
def interAminoFeature(file):
    #read protein sequences from files 
    for filepath in file:
        content = pd.read_excel(filepath)
        protein = content["pro"].values
        work_book = load_workbook(filepath)
        sheet = work_book.active
        amino_name = ["A", "C", "D", "E", "F", "G", "H", "I", "K", "L", "M", "N", "P", "Q", "R", "S", "T", "V", "W", "Y"]
        col_name = []
        for item in amino_name:
            col_name.append("F" + item)
            col_name.append("A" + item)
            col_name.append("D" + item)
            col_name.append("M" + item)
            col_name.append("m" + item)
        for i in range(len(col_name)):
            sheet.insert_cols(4+i)
        for index in range(len(col_name)):
            sheet.cell(row=1,column=4+index).value = col_name[index]
        row = 2
        #start to perform feature calculation,and save data
        for num in range(len(protein)):
            feature = []
            seq = list(protein[num].upper())
            length = len(seq)
            amino = ["A","C","D","E","F","G","H","I","K","L","M","N","P","Q","R","S","T","V","W","Y"]
            index_arr = []
            for item in amino:
                index =  [i for i, x in enumerate(seq) if x == item]
                index_arr.append(index)
            distance = []
            distance_order = []
            print(index_arr)
            for i in range(20):
                dis = []
                for j in range(len(index_arr[i])-1):
                    dis.append(index_arr[i][j+1]-index_arr[i][j])
                
                distance.append(dis)
                distance_order.append(sorted(dis))
            for i in range(20):
                item = distance_order[i]
                feature.append((len(index_arr[i]))/length)
                feature.append(np.mean(item))
                feature.append(np.var(item))
                if len(item)==0:
                    feature.append(0)
                else:
                    occur_time = Counter(item)  
                    print(occur_time)
                    mode = max(occur_time.values())
                    mode_item = []
                    for (key, value) in occur_time.items():
                        if value == mode:
                            mode_item.append(key) 
                    feature.append(min(mode_item))
                feature.append(np.median(item))
            for i in range(len(feature)):
                sheet.cell(row = row,column= 4+i).value = feature[i]
            row += 1
        work_book.save(filepath)