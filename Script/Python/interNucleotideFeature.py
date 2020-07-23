import rpy2.robjects as robjects
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from rpy2.robjects.packages import importr
import os
import pandas as pd
import re
import numpy as np
import math
import scipy
from openpyxl.utils import get_column_letter
from collections import Counter
from scipy.fftpack import fft

#Calculate the inter-nucleotide Features
def interNucleotideFeature(file):
    # Read gene sequences from files
    for filepath in file:
        print(filepath)
        content = pd.read_excel(filepath)
        gene = content["gene"].values
        col_name = ["Fa", "Aa", "Da", "Ma", "ma", "Fg", "Ag", "Dg", "Mg", "mg", "Fc", "Ac", "Dc","Mc", "mc", "Ft", "At", "Dt", "Mt", "mt"]
        work_book = load_workbook(filepath)
        sheet = work_book.active
        for i in range(len(col_name)):
            sheet.insert_cols(4+i)
        for index in range(len(col_name)):
            sheet.cell(row=1,column=4+index).value = col_name[index]
        row = 2
        #start to perform feature calculation,and save data
        for num in range(len(gene)):
            feature = []
            seq = list(gene[num])
            length = len(seq)
            indexA = [i for i, x in enumerate(seq) if x == 'A' or x=="a"]
            indexG = [i for i, x in enumerate(seq) if x == 'G' or x == "g"]
            indexC = [i for i, x in enumerate(seq) if x == 'C' or x == "c"]
            indexT = [i for i, x in enumerate(seq) if x == 'T' or x == "t"]
            index_arr = [indexA,indexG,indexC,indexT]
            distance = []
            distance_order = []
            print(index_arr)
            for i in range(4):
                dis = []
                for j in range(len(index_arr[i])-1):
                    dis.append(index_arr[i][j+1]-index_arr[i][j])
                distance.append(dis)
                distance_order.append(sorted(dis))
            for i in range(4):
                item = distance_order[i]
                feature.append((len(index_arr[i]))/length)
                feature.append(np.mean(item))
                feature.append(np.var(item))
                if len(item)==0:
                    feature.append(0)
                else:
                    occur_time = Counter(item)  
                    print(occur_time)
                    mode = max(occur_time.values())
                    mode_item = []
                    for (key, value) in occur_time.items():
                        if value == mode:
                            mode_item.append(key) 
                    feature.append(min(mode_item))
                feature.append(np.median(item))
            for i in range(len(feature)):
                sheet.cell(row=row,column=4+i).value = feature[i]
            row += 1
        work_book.save(filepath)