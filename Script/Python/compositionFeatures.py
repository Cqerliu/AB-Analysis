import rpy2.robjects as robjects
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from rpy2.robjects.packages import importr
import os
import pandas as pd
import re
import numpy as np
import math
import scipy
from openpyxl.utils import get_column_letter
from collections import Counter
from scipy.fftpack import fft

#Calculate the composition features
def compositionFeatures(file):
    #read sequences from files
    for filepath in file:
        content = pd.read_excel(filepath)
        protein = content["pro"].values
        name = ["A1","A2","A3","C1","C2","C3","U1","U2","U3","G1","G2","G3"]
        work_book = load_workbook(filepath)
        sheet = work_book.active
        sheet.insert_cols(4)
        sheet.insert_cols(5)
        sheet.insert_cols(6)
        sheet.insert_cols(7)
        sheet.insert_cols(8)
        sheet.insert_cols(9)
        sheet.insert_cols(10)
        sheet.insert_cols(11)
        sheet.insert_cols(12)
        sheet.insert_cols(13)
        sheet.insert_cols(14)
        sheet.insert_cols(15)
        for index in range(len(name)):
            sheet.cell(row=1,column=4+index).value = name[index]
        row = 2
        #start to perform feature calculation,and save data
        for num in range(len(protein)):
            A = protein[num].count("A")
            R = protein[num].count("R")
            D = protein[num].count("D")
            C = protein[num].count("C")
            Q = protein[num].count("Q")
            E = protein[num].count("E")
            G = protein[num].count("G")
            H = protein[num].count("H")
            I = protein[num].count("I")
            N = protein[num].count("N")
            L = protein[num].count("L")
            K = protein[num].count("K")
            M = protein[num].count("M")
            F = protein[num].count("F")
            P = protein[num].count("P")
            S = protein[num].count("S")
            T = protein[num].count("T")
            W = protein[num].count("W")
            Y = protein[num].count("Y")
            V = protein[num].count("V")
            pro_len = len(protein[num])
            A1 =  (I+M+T+N+K+S+R)/pro_len
            A2 = (Y+H+Q+N+K+D+E)/pro_len
            A3 = (L+S+P+Q+R+I+T+K+V+A+E+G)/pro_len
            C1 = (L+P+H+R+Q)/pro_len
            C2 = (S+P+T+A)/pro_len
            C3 = (F+S+Y+C+L+P+H+R+I+T+N+S+V+A+D+G)/pro_len
            U1 = sum([F,L,S,Y,C,W])/pro_len
            U2 = sum([F,L,I,M,V])/pro_len
            U3 = sum([ F,S,Y,C,L,P,H,R,I,T,N,S,V,A,D,G])/pro_len
            G1 = sum([V,A,D,E,G])/pro_len
            G2 = sum([C,W,R,S,G])/pro_len
            G3 = sum([L,S,W,P,Q,M,T,K,R,V,A,E,G])/pro_len
            feature = [A1,A2,A3,C1,C2,C3,U1,U2,U3,G1,G2,G3]
            for index in range(4,16):
                sheet.cell(row = row,column=index).value = feature[index-4]
            row = row +1
        work_book.save(filepath)