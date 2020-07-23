import rpy2.robjects as robjects
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from rpy2.robjects.packages import importr
import os
import pandas as pd
import re
import numpy as np
import math
import scipy
from openpyxl.utils import get_column_letter
from collections import Counter
from scipy.fftpack import fft
#Calculate the hurst index of the sequence
def hurst(file):
    fractal = importr('fractal')
	# Read gene sequences from files
    for filepath in file:
        print(filepath)
        work_book = load_workbook(filepath)
        sheet = work_book.active
        sheet.cell(row=1,column=4).value = "hurst"
        content = pd.read_excel(filepath)
        gene = content["gene"]
        work_book = load_workbook(filepath)
        sheet = work_book.active
        sheet.cell(row=1,column=5).value = "hurst"
        row = 2
        col = 5
        #start to perform feature calculation,and save data
        for num  in range(len(gene)):
            print(num)
            seq = gene[num].replace("A","0")
            seq = seq.replace("G", "1")
            seq = seq.replace("C", "2")
            seq = seq.replace("T", "3")
            seq = seq.replace("B", "3")
            seq = seq.replace("D", "1")
            seq = seq.replace("H", "0")
            seq = seq.replace("V", "2")
            seq = seq.replace("M", "0")
            seq = seq.replace("R", "1")
            seq = seq.replace("K", "1")
            seq = seq.replace("S", "2")
            seq = seq.replace("W", "3")
            seq = seq.replace("Y", "2")
            seq = seq.replace("N", "2")
            seq = seq.replace("I", "4")
            seq = list(map(int, seq))
            robjects.r.source("cal_hurst.R")
            sequence = robjects.IntVector(seq)
            hurst_value = robjects.r.calHurst(sequence)
            sheet.cell(row=row,column=col).value = list(hurst_value)[0]
            row += 1
        work_book.save(filepath)