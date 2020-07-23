import rpy2.robjects as robjects
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from rpy2.robjects.packages import importr
import os
import pandas as pd
import re
import numpy as np
import math
import scipy
from openpyxl.utils import get_column_letter
from collections import Counter
from scipy.fftpack import fft

#Calculate amino abscid usage features
def aminoAcidUsageFeatures(file):
    #Read protein sequences from files, this part is not necessary
    for filepath in file:
        content = pd.read_excel(filepath)
        protein = content["pro"]
        work_book = load_workbook(filepath)
        sheet = work_book.active
        name = ["A","R","D","C","Q","E","G","H","I","N","L","K","M","F","S","T","W","Y","V","rare_aa_ratio","close_aa_ratio"]
        col = 4
        for item in name:
            sheet.cell(1,col).value = item
            col += 1
        row = 2
        print(len(protein))
        #start to perform feature calculation
        for num in range(len(protein)):
            A = protein[num].count("A")/len(protein[num])
            R = protein[num].count("R")/len(protein[num])
            D = protein[num].count("D")/len(protein[num])
            C = protein[num].count("C")/len(protein[num])
            Q = protein[num].count("Q")/len(protein[num])
            E = protein[num].count("E")/len(protein[num])
            G = protein[num].count("G")/len(protein[num])
            H = protein[num].count("H")/len(protein[num])
            I = protein[num].count("I")/len(protein[num])
            N = protein[num].count("N")/len(protein[num])
            L = protein[num].count("L")/len(protein[num])
            K = protein[num].count("K")/len(protein[num])
            M = protein[num].count("M")/len(protein[num])
            F = protein[num].count("F")/len(protein[num])
            P = protein[num].count("P")/len(protein[num])
            S = protein[num].count("S")/len(protein[num])
            T = protein[num].count("T")/len(protein[num])
            W = protein[num].count("W")/len(protein[num])
            Y = protein[num].count("Y")/len(protein[num])
            V = protein[num].count("V")/len(protein[num])
            Rare_aa_ratio = (H+M+W)/len(protein[num])
            close_aa_ratio = (C+W+Y)/len(protein[num])
            feature = [A,R,D,C,Q,E,G,H,I,N,L,K,M,F,P,S,T,W,Y,V,Rare_aa_ratio,close_aa_ratio]
            # Save data
            col = 4
            for item in feature:
                sheet.cell(row,col).value = item
                col = col +1
            row = row + 1
        work_book.save(filepath)