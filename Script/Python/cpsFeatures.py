import rpy2.robjects as robjects
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from rpy2.robjects.packages import importr
import os
import pandas as pd
import re
import numpy as np
import math
import scipy
from openpyxl.utils import get_column_letter
from collections import Counter
from scipy.fftpack import fft

##Calculate the Cumulative Fourier power spectrum features
def cpsFeatures(file):
    #read sequences from files
    for filepath in file:
        content = pd.read_excel(filepath)
        gene = content["gene"].values
        work_book = load_workbook(filepath)
        sheet = work_book.active
        col_name = ["m1a", "m2a", "m1g", "m2g", "m1c", "m2c", "m1t", "m2t", "cm1a", "cm2a", "cm1g", "cm2g", "cm1c", "cm2c", "cm1t", "cm2t"]
        for i in range(len(col_name)):
            sheet.insert_cols(4+i)
        for index in range(len(col_name)):
            sheet.cell(row=1,column=4+index).value = col_name[index]
        row = 2
        len_gene = len(gene)
        #start to perform feature calculation,and save data
        for num in range(len_gene):
            na = 0
            nc = 0
            ng = 0
            nt = 0
            ua = []
            ug = []
            uc = []
            ut = []
            seq = gene[num]
            seq_len = len(seq)
            for i in range(seq_len):
                if seq[i] == "A" or seq[i] == "a":
                    ua.append(1)
                    ug.append(0)
                    uc.append(0)
                    ut.append(0)
                    na += 1
                elif seq[i] == "G" or seq[i] == "g":
                    ua.append(0)
                    ug.append(1)
                    uc.append(0)
                    ut.append(0)
                    ng += 1
                elif seq[i] == "C" or seq[i] == "c":
                    ua.append(0)
                    ug.append(0)
                    uc.append(1)
                    ut.append(0)
                    nc += 1
                elif seq[i] == "T" or seq[i] == "t":
                    ua.append(0)
                    ug.append(0)
                    uc.append(0)
                    ut.append(1)
                    nt += 1
                else:
                    ua.append(0)
                    ug.append(0)
                    uc.append(0)
                    ut.append(0)
            UA = fft(ua)  
            UG = fft(ug)
            UC = fft(uc)
            UT = fft(ut)
            PSA = [0 for _ in range(seq_len)]
            PSG = [0 for _ in range(seq_len)]
            PSC = [0 for _ in range(seq_len)]
            PST = [0 for _ in range(seq_len)]
            for i in range(seq_len):
                PSA[i] = pow(abs(UA[i]), 2)
                PSG[i] = pow(abs(UG[i]), 2)
                PSC[i] = pow(abs(UC[i]), 2)
                PST[i] = pow(abs(UT[i]), 2)
            CPSA = [0 for _ in range(seq_len)]  
            CPSG = [0 for _ in range(seq_len)]
            CPSC = [0 for _ in range(seq_len)]
            CPST = [0 for _ in range(seq_len)]
            for i in range(1, seq_len):
                CPSA[i] = CPSA[i - 1] + PSA[i]
                CPSG[i] = CPSG[i - 1] + PSG[i]
                CPSC[i] = CPSC[i - 1] + PSC[i]
                CPST[i] = CPST[i - 1] + PST[i]
            del CPSA[0]
            del CPSG[0]
            del CPSC[0]
            del CPST[0]
            MA1 = sum(CPSA) / seq_len
            MA2 = sum([num * num for num in CPSA]) / (na * (seq_len - na) * seq_len * seq_len)
            MG1 = sum(CPSG) / seq_len
            MG2 = sum([num * num for num in CPSG]) / (ng * (seq_len - ng) * seq_len * seq_len)
            MC1 = sum(CPSC) / seq_len
            MC2 = sum([num * num for num in CPSC]) / (nc * (seq_len - nc) * seq_len * seq_len)
            MT1 = sum(CPST) / seq_len
            MT2 = sum([num * num for num in CPST]) / (nt * (seq_len - nt) * seq_len * seq_len)
            CMA1 = sum([abs(num) for num in (CPSA - np.mean(CPSA))]) / seq_len
            CMA2 = sum([num * num for num in (CPSA - np.mean(CPSA))]) / (na * (seq_len - na) * seq_len * seq_len)
            CMG1 = sum([abs(num) for num in (CPSG - np.mean(CPSG))]) / seq_len
            CMG2 = sum([num * num for num in (CPSG - np.mean(CPSG))]) / (ng * (seq_len - ng) * seq_len * seq_len)
            CMC1 = sum([abs(num) for num in (CPSC - np.mean(CPSC))]) / seq_len
            CMC2 = sum([num * num for num in (CPSC - np.mean(CPSC))]) / (nc * (seq_len - nc) * seq_len * seq_len)
            CMT1 = sum([abs(num) for num in (CPST - np.mean(CPST))]) / seq_len
            CMT2 = sum([num * num for num in (CPST - np.mean(CPST))]) / (nt * (seq_len - nt) * seq_len * seq_len)
            res = [ MA1, MA2, MG1, MG2, MC1, MC2, MT1, MT2, CMA1, CMA2, CMG1, CMG2,CMC1, CMC2, CMT1, CMT2]
            for i in range(len(res)):
                sheet.cell(row=row, column=4 + i).value = res[i]
            row += 1
        work_book.save(filepath)